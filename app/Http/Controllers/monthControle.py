from flask import Flask, request, send_file, after_this_request
import os
import tempfile
import win32com.client
import pythoncom
from openpyxl import load_workbook
import time
import threading
import datetime
import os
import pythoncom
import win32com.client as win32
import hashlib
def convert_xls_to_xlsx(xls_path):
    # إنشاء مسار الملف الجديد
    base, _ = os.path.splitext(xls_path)
    xlsx_path = base + ".xlsx"

    pythoncom.CoInitialize()  # تهيئة COM
    excel = win32.DispatchEx("Excel.Application")
    excel.DisplayAlerts = False  # منع الرسائل المنبثقة

    try:
        wb = excel.Workbooks.Open(xls_path)
        # حفظ كـ XLSX (الرقم 51 = صيغة Excel 2007+)
        wb.SaveAs(xlsx_path, FileFormat=51)
        wb.Close()
    finally:
        excel.Quit()

    return xlsx_path

app = Flask(__name__)


def make_code(input_number: int, salt: str = "my_secret_salt") -> str:
    # حوّل الإدخال لنص وثبّت معه ملح (salt) لزيادة الأمان/التنوع
    s = f"{salt}:{input_number}"
    h = hashlib.sha256(s.encode()).hexdigest()  # تجزئة ثابتة
    # نقسم التجزئة لخمسة أجزاء صغيرة ونحوّل كل جزء لرقم 0-9
    digits = []
    for i in range(5):
        chunk = h[i*12:(i+1)*12]  # 12 hex chars لكل جزء
        n = int(chunk, 16)
        digits.append(str(n % 10))
    return "".join(digits)


@app.route("/download")
def download():
    file_path = request.args.get("file_path")
    account = request.args.get("account")
    name_type = request.args.get("type")
    sub_type = request.args.get("subType")
    password = request.args.get("password")
    print(make_code(account))
    print(password)
    if (password) == str(make_code(account)):
        print('sees')
    

    print(name_type)

    if not file_path or not account or not name_type:
        return "❌ المعطيات ناقصة", 400

    if not os.path.exists(file_path):
        return f"❌ الملف غير موجود: {file_path}", 404

    if name_type == 'استاذة':
            # استخراج الاسم من ملف Excel
        try:
            wb_openpyxl = load_workbook(file_path, data_only=True)
            sheet = wb_openpyxl['SOURCE-NOUVE (2)']
            search_col_index = 13  # العمود M
            name_col_index = 6     # العمود F
            name = None

            for row in range(1, sheet.max_row + 1):
                val = sheet.cell(row=row, column=search_col_index).value
                if str(val).strip() == str(account).strip():
                    name = sheet.cell(row=row, column=name_col_index).value
                    break

            wb_openpyxl.close()
        except Exception as e:
            return f"❌ خطأ في قراءة Excel: {e}", 500

        if not name:
            return "❌ لم يتم العثور على الاسم", 404

        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        try:
            # محاولة ضبط الحساب التلقائي على مستوى التطبيق
            try:
                excel.Calculation = -4105  # xlCalculationAutomatic
            except Exception as calc_err:
                print(f"⚠️ تعذر ضبط الحساب التلقائي على مستوى التطبيق: {calc_err}")

            wb = excel.Workbooks.Open(file_path, UpdateLinks=True)
            ws_target = wb.Sheets("كشف الراتب عربية")

            # تفعيل الحساب لكل ورقة وتحديثها
            for ws in wb.Sheets:
                try:
                    ws.EnableCalculation = True
                    ws.Calculate()
                except Exception:
                    pass

            # كتابة الاسم في I1
            ws_target.Range("I1").Value = name
            ws_target.Range("F43").Value = datetime.datetime.now().strftime("%d-%m-%Y")

            print(f"📌 تم تغيير I1 إلى: {name}")

            # تحديث جميع PivotTables
            for sheet in wb.Sheets:
                try:
                    for pivot in sheet.PivotTables():
                        pivot.RefreshTable()
                except Exception:
                    pass

            # تحديث جميع الروابط وPower Query
            try:
                wb.RefreshAll()
            except Exception:
                pass

            # إعادة الحساب الشاملة
            excel.CalculateFullRebuild()

            # حفظ القيم القديمة
            old_values = [cell.Value for cell in ws_target.Range("A1:H46")]

            # انتظار التغيير في القيم
            max_tries = 30
            for i in range(max_tries):
                new_values = [cell.Value for cell in ws_target.Range("A1:H46")]
                if new_values != old_values:
                    print("✅ تم تحديث القيم في منطقة الطباعة")
                    break
                print(f"⏳ محاولة {i+1}: لم تتغير القيم بعد")
                time.sleep(1)
            else:
                print("⚠️ لم تتغير القيم بعد 30 محاولة")

            # إعدادات الطباعة
            ws_target.PageSetup.PrintArea = "A1:H46"
            ws_target.PageSetup.Orientation = 1
            ws_target.PageSetup.Zoom = False
            ws_target.PageSetup.FitToPagesWide = 1
            ws_target.PageSetup.FitToPagesTall = 1

            # تصدير PDF مؤقت
            tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            pdf_path = tmp_pdf.name
            tmp_pdf.close()
            ws_target.ExportAsFixedFormat(0, pdf_path)

            wb.Close(False)
            excel.Quit()

        except Exception as e:
            excel.Quit()
            return f"❌ خطأ أثناء التصدير: {e}", 500

        @after_this_request
        def remove_file(response):
            def delete_later(path):
                try:
                    os.remove(path)
                    print("🗑️ تم حذف PDF بعد الإرسال")
                except Exception as e:
                    print(f"⚠️ فشل حذف PDF لاحقًا: {e}")
            threading.Timer(5.0, delete_later, args=[pdf_path]).start()
            return response

        return send_file(pdf_path, as_attachment=True)
    
    
    if name_type == 'عامل متقاعد':
                  # استخراج الاسم من ملف Excel
        try:
            xlsx_file = convert_xls_to_xlsx(file_path)

            wb_openpyxl = load_workbook(xlsx_file, data_only=True)
            sheet = wb_openpyxl['source']
            search_col_index = 4  # العمود M
            name_col_index = 1     # العمود F
            name = None

            for row in range(1, sheet.max_row + 1):
                val = sheet.cell(row=row, column=search_col_index).value
                if str(val).strip() == str(account).strip():
                    name = sheet.cell(row=row, column=name_col_index).value
                    break

            wb_openpyxl.close()
        except Exception as e:
            return f"❌ خطأ في قراءة Excel: {e}", 500

        if not name:
            return "❌ لم يتم العثور على الاسم", 404

        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        try:
            # محاولة ضبط الحساب التلقائي على مستوى التطبيق
            try:
                excel.Calculation = -4105  # xlCalculationAutomatic
            except Exception as calc_err:
                print(f"⚠️ تعذر ضبط الحساب التلقائي على مستوى التطبيق: {calc_err}")

            wb = excel.Workbooks.Open(file_path, UpdateLinks=True)
            ws_target = wb.Sheets("fiche")

            # تفعيل الحساب لكل ورقة وتحديثها
            for ws in wb.Sheets:
                try:
                    ws.EnableCalculation = True
                    ws.Calculate()
                except Exception:
                    pass

            if str(name).strip() == "1":
                ws_target.Range("A2").Value = name
            else:
                try:
                    ws_target.Range("A2").Value = int(name) - 1
                except ValueError:
                    # إذا لم يكن رقم، نكتفي بوضعه كما هو
                    ws_target.Range("A2").Value = name


            

            print(f"📌 تم تغيير I1 إلى: {name}")

            # تحديث جميع PivotTables
            for sheet in wb.Sheets:
                try:
                    for pivot in sheet.PivotTables():
                        pivot.RefreshTable()
                except Exception:
                    pass

            # تحديث جميع الروابط وPower Query
            try:
                wb.RefreshAll()
            except Exception:
                pass

            # إعادة الحساب الشاملة
            excel.CalculateFullRebuild()

            # حفظ القيم القديمة
            old_values = [cell.Value for cell in ws_target.Range("A1:H46")]

            # انتظار التغيير في القيم
            max_tries = 30
            for i in range(max_tries):
                new_values = [cell.Value for cell in ws_target.Range("A1:H46")]
                if new_values != old_values:
                    print("✅ تم تحديث القيم في منطقة الطباعة")
                    break
                print(f"⏳ محاولة {i+1}: لم تتغير القيم بعد")
                time.sleep(1)
            else:
                print("⚠️ لم تتغير القيم بعد 30 محاولة")

            # إعدادات الطباعة
            ws_target.PageSetup.PrintArea = "A1:H38"
            ws_target.PageSetup.Orientation = 1
            ws_target.PageSetup.Zoom = False
            ws_target.PageSetup.FitToPagesWide = 1
            ws_target.PageSetup.FitToPagesTall = 1

            # تصدير PDF مؤقت
            tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            pdf_path = tmp_pdf.name
            tmp_pdf.close()
            ws_target.ExportAsFixedFormat(0, pdf_path)

            wb.Close(False)
            excel.Quit()

        except Exception as e:
            excel.Quit()
            return f"❌ خطأ أثناء التصدير: {e}", 500

        @after_this_request
        def remove_file(response):
            def delete_later(path):
                try:
                    os.remove(path)
                    print("🗑️ تم حذف PDF بعد الإرسال")
                except Exception as e:
                    print(f"⚠️ فشل حذف PDF لاحقًا: {e}")
            threading.Timer(5.0, delete_later, args=[pdf_path]).start()
            return response

        return send_file(pdf_path, as_attachment=True)

    if name_type == "عامل دائم":
        print("sub type = " + sub_type + " with type = " + str(type(sub_type)))
         # استخراج الاسم من ملف Excel
        try:
            

            wb_openpyxl = load_workbook(file_path, data_only=True)
            sheet = wb_openpyxl['sours']

            search_col_index = 5  # العمود M
            found_row_number = None

            for row in range(1, sheet.max_row + 1):
                val = sheet.cell(row=row, column=search_col_index).value
                if str(val).strip() == str(account).strip():
                    found_row_number = row  # نحفظ رقم الصف
                    break

            wb_openpyxl.close()

            print("✅ رقم الصف:", found_row_number)

        except Exception as e:
            return f"❌ خطأ في قراءة Excel: {e}", 500

        if not found_row_number:
            return "❌ لم يتم العثور على الاسم", 404

        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            # محاولة ضبط الحساب التلقائي على مستوى التطبيق
            try:
                excel.Calculation = -4105  # xlCalculationAutomatic
            except Exception as calc_err:
                print(f"⚠️ تعذر ضبط الحساب التلقائي على مستوى التطبيق: {calc_err}")

            wb = excel.Workbooks.Open(file_path, UpdateLinks=True)
            if sub_type in ["الأسلاك المشتركة", "سلك المهنيين وسائقي السيارات والحجاب", "2","3"]:
                ws_target = wb.Sheets("fiche_COMM_ouv")
                print("✅ matched fiche_COMM_ouv")

            elif sub_type in ["سلك الموظفين المنتمين للأسلاك التعليم العالي", "4"] :
                ws_target = wb.Sheets("fiche_Ens_Sup")
                print("✅ matched fiche_Ens_Sup")

            elif sub_type in [
                "أسلاك تقنية خاصة خاصة بالإدارة السكن والعمران",
                "أسلاك شبه الطبيين للصحة العمومية",
                "سلك الممارسين الطبيين العامين للصحة العمومية","5","6","7"
            ]:
                ws_target = wb.Sheets("fiche_Habit_sant")
                if( found_row_number !=5):
                    found_row_number = found_row_number -2
                print("✅ matched fiche_Habit_sant")

            elif sub_type in ["بيطرة" , "8"]:
                ws_target = wb.Sheets("fiche_Veteriarian")
                print("✅ matched fiche_Veteriarian")
                found_row_number =found_row_number -2


            else:
                return f"❌ لا يوجد ورقة للنوع: {sub_type}", 400

            print(f"📄 Working on sheet: {ws_target.Name}")


            # تفعيل الحساب لكل ورقة وتحديثها
            for ws in wb.Sheets:
                try:
                    ws.EnableCalculation = True
                    ws.Calculate()
                except Exception:
                    pass

            
            ws_target.Range("A2").Value = found_row_number

           
            

            print(f"📌 تم تغيير a2 إلى: {found_row_number}")

            # تحديث جميع PivotTables
            for sheet in wb.Sheets:
                try:
                    for pivot in sheet.PivotTables():
                        pivot.RefreshTable()
                except Exception:
                    pass

            # تحديث جميع الروابط وPower Query
            try:
                wb.RefreshAll()
            except Exception:
                pass

            # إعادة الحساب الشاملة
            excel.CalculateFullRebuild()

            # حفظ القيم القديمة
            old_values = [cell.Value for cell in ws_target.Range("A1:H46")]

            # انتظار التغيير في القيم
            max_tries = 30
            for i in range(max_tries):
                new_values = [cell.Value for cell in ws_target.Range("A1:H46")]
                if new_values != old_values:
                    print("✅ تم تحديث القيم في منطقة الطباعة")
                    break
                print(f"⏳ محاولة {i+1}: لم تتغير القيم بعد")
                time.sleep(1)
            else:
                print("⚠️ لم تتغير القيم بعد 30 محاولة")

            
            if sub_type in ["الأسلاك المشتركة", "سلك المهنيين وسائقي السيارات والحجاب", "2","3"]:
               if(ws_target.Range("K36").Value != 0):
                    return f"❌ خطأ في البيانات: {e}", 500
              
            elif sub_type in ["سلك الموظفين المنتمين للأسلاك التعليم العالي", "4"] :
                if(ws_target.Range("K39").Value != 0):
                    return f"❌ خطأ في البيانات: {e}", 500

            elif sub_type in [
                "أسلاك تقنية خاصة خاصة بالإدارة السكن والعمران",
                "أسلاك شبه الطبيين للصحة العمومية",
                "سلك الممارسين الطبيين العامين للصحة العمومية","5","6","7","بيطرة" , "8"
            ]:
                if(ws_target.Range("K38").Value != 0):
                    return f"❌ خطأ في البيانات: {e}", 500


            # إعدادات الطباعة
            ws_target.PageSetup.PrintArea = "A1:I48"
            ws_target.PageSetup.Orientation = 1
            ws_target.PageSetup.Zoom = False
            ws_target.PageSetup.FitToPagesWide = 1
            ws_target.PageSetup.FitToPagesTall = 1

            # تصدير PDF مؤقت
            tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            pdf_path = tmp_pdf.name
            tmp_pdf.close()
            ws_target.ExportAsFixedFormat(0, pdf_path)

            wb.Close(False)
            excel.Quit()

        except Exception as e:
            excel.Quit()
            return f"❌ خطأ في البيانات: {e}", 500

        @after_this_request
        def remove_file(response):
            def delete_later(path):
                try:
                    os.remove(path)
                    print("🗑️ تم حذف PDF بعد الإرسال")
                except Exception as e:
                    print(f"⚠️ فشل حذف PDF لاحقًا: {e}")
            threading.Timer(5.0, delete_later, args=[pdf_path]).start()
            return response

        return send_file(pdf_path, as_attachment=True)
    return "❌ النوع غير مدعوم", 400

if __name__ == "__main__":
    app.run(  port=5000, debug=True)
